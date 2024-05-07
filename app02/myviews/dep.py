from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook

from app02.models import Department


def dep_list(request):
    # 检查登录情况
    if not request.session.get('info'):
        return redirect('/login/')
    dep_set = Department.objects.all()

    return render(request, 'dep_list.html', {'dep_set': dep_set})


def dep_add(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        Department.objects.create(title=title)
        return redirect('/dep/list')

    if request.method == 'GET':
        return render(request, 'dep_add.html')


def dep_del(request):
    id = request.GET.get('id')
    Department.objects.filter(id=id).delete()

    return redirect('/dep/list')


def dep_edit(request, id):
    if request.method == 'POST':
        title = request.POST.get('title')
        Department.objects.filter(id=id).update(title=title)
        return redirect('/dep/list')
    dep_data = Department.objects.filter(id=id).first()
    return render(request, 'dep_edit.html', {'dep_data': dep_data})


@csrf_exempt
def dep_upload(request):
    file_obj = request.FILES.get('avatar')
    print(type(file_obj), file_obj)

    if file_obj is None:  # 未上传文件
        return redirect('/dep/list')

    wb = load_workbook(file_obj)
    ws = wb.worksheets[0]
    for row in ws.iter_rows(min_row=2):
        dep = row[0].value
        print(dep)
        if not Department.objects.filter(title=dep).exists():
            Department.objects.create(title=dep)

    return redirect('/dep/list')
